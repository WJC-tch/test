#coding:UTF-8
import openpyxl
import time

path = r'C:\Users\jattc\Desktop\Gift4TimeSheet\TSE_FOLDERNew\TSE_900238.xlsm'
#path = r'C:\Users\jattc\Desktop\Gift4TimeSheet\TSE_FOLDERNew\111.xlsm'
excel_write = openpyxl.load_workbook(path,keep_vba=True)

table_write = excel_write['TimeSheet']
print(table_write.cell(2,2).value)
c = table_write.cell(row=142, column=24)
#table_write.cell(142, 24).value= time.strftime("%H:%M", time.localtime())
table_write.cell(142, 24).value='9:30'
print(c.value)
excel_write.save('TSE.xlsm')